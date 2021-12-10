from django.shortcuts import redirect, render
import openpyxl
from .credentials import get_access


data=get_access("14xGox9yPdCtGrOpgQszVPIKIcy2K24cNkkjci7mheUg")
def excel_update(request,worksheet):
        excel_data = list()
        if list(worksheet.iter_rows(1,1,None,None,1))!=[('paperid', 'authorcount', 'name', 'journal', 'paper', 'edition', 'year', 'spage', 'epage')]:
             return render(request,"excel.html",{"error":"Not A Valid File"})
        else:
            for row in worksheet.iter_rows(2,None,None,None,1):
                print(row)
                data.append_row(row)
            return redirect("home")


def excel(request):
     if(request.POST.get('submit')):
        excel_file = request.FILES["excel"]
        try:
            wb = openpyxl.load_workbook(excel_file)
        except:
            return render(request,"excel.html",{"error":"Select a Valid File As Given Below"})
        sheets = wb.sheetnames
        worksheet = wb[sheets[0]]
        return excel_update(request,worksheet)
     return render(request,"excel.html")